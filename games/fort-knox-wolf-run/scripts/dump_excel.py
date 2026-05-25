#!/usr/bin/env python3
"""Fort Knox Wolf Run — raw Excel dumper.

The original `PAR_Sheets_FortKnoxWolfRun.xlsx` has a malformed `textRotation`
attribute in its stylesheet (vendor-export bug). We unzip → patch → reload
before dumping. Output schema mirrors `games/ce-copy-test/raw/*` so the
downstream IR parser can reuse helpers.
"""
import json, os, re, shutil, subprocess, tempfile, zipfile
from pathlib import Path
import openpyxl

SRC_RAW = Path("/Users/vanvinklstudio/Desktop/Bojan/PAR_Sheets_FortKnoxWolfRun.xlsx")
OUT = Path(__file__).resolve().parent.parent / "raw"
OUT.mkdir(parents=True, exist_ok=True)

def patch_xlsx(src: Path) -> Path:
    tmp_zip = Path(tempfile.mkdtemp()) / "fixed.xlsx"
    tmp_dir = tmp_zip.parent / "unpack"
    with zipfile.ZipFile(src, 'r') as z:
        z.extractall(tmp_dir)
    styles = tmp_dir / "xl" / "styles.xml"
    s = styles.read_text()
    s2 = re.sub(r'textRotation="(\d+)"',
                lambda m: '' if int(m.group(1)) > 180 else m.group(0), s)
    styles.write_text(s2)
    # rezip
    with zipfile.ZipFile(tmp_zip, 'w', zipfile.ZIP_DEFLATED) as z:
        for p in tmp_dir.rglob('*'):
            if p.is_file():
                z.write(p, p.relative_to(tmp_dir))
    return tmp_zip

def cell_value(v):
    if v is None: return None
    if isinstance(v, (int, float, str, bool)): return v
    return str(v)

def dump(wb, suffix):
    for sn in wb.sheetnames:
        ws = wb[sn]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                v = cell_value(c.value)
                if v is None or (isinstance(v, str) and v == ""): continue
                cells[c.coordinate] = v
        path = OUT / f"{sn}.{suffix}.json"
        path.write_text(json.dumps(cells, ensure_ascii=False, indent=2))
        print(f"  {sn}.{suffix}.json :: {len(cells)} cells")

def tsv(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append("\t".join("" if v is None else str(v) for v in r))
        (OUT / f"{sn}.tsv").write_text("\n".join(rows))
        print(f"  {sn}.tsv :: {len(rows)} rows")

def main():
    print(f"[fort-knox-wolf-run] dumping {SRC_RAW}")
    fixed = patch_xlsx(SRC_RAW)
    wb_v = openpyxl.load_workbook(fixed, data_only=True)
    wb_f = openpyxl.load_workbook(fixed, data_only=False)
    print("[values]")
    dump(wb_v, "cells")
    print("[formulas]")
    dump(wb_f, "formulas")
    print("[tsv]")
    tsv(wb_v)
    print(f"\nDone → {OUT}")

if __name__ == "__main__":
    main()
