#!/usr/bin/env python3
"""CE COPY TEST — RAW Excel dumper.

Dumps every cell from every sheet of ParSheets_CashEruption 1.xlsx into:
  - raw/<sheet>.cells.json   — addr → value (data_only=True, evaluated formulas)
  - raw/<sheet>.formulas.json — addr → formula text (data_only=False)
  - raw/<sheet>.tsv          — human-readable grid

This is the single source of truth. Every downstream artefact (IR JSON,
Rust engine, TS engine, PAR report) must round-trip through these dumps
bit-identical.
"""
from __future__ import annotations
import json
from pathlib import Path
import openpyxl

SRC = Path("/Users/vanvinklstudio/Desktop/Bojan/ParSheets_CashEruption 1.xlsx")
OUT = Path(__file__).resolve().parent.parent / "raw"
OUT.mkdir(parents=True, exist_ok=True)


def cell_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    # datetimes, etc — stringify
    return str(v)


def dump(workbook, data_only: bool, suffix: str):
    for sn in workbook.sheetnames:
        ws = workbook[sn]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                v = cell_value(c.value)
                if v is None or (isinstance(v, str) and v == ""):
                    continue
                cells[c.coordinate] = v
        path = OUT / f"{sn}.{suffix}.json"
        path.write_text(json.dumps(cells, ensure_ascii=False, indent=2))
        print(f"  {sn}.{suffix}.json :: {len(cells)} non-empty cells")


def tsv(workbook):
    for sn in workbook.sheetnames:
        ws = workbook[sn]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append("\t".join("" if v is None else str(v) for v in r))
        path = OUT / f"{sn}.tsv"
        path.write_text("\n".join(rows))
        print(f"  {sn}.tsv :: {len(rows)} rows")


def main():
    print(f"[ce-copy-test] dumping {SRC}")
    wb_vals = openpyxl.load_workbook(SRC, data_only=True)
    wb_form = openpyxl.load_workbook(SRC, data_only=False)
    print("[values]")
    dump(wb_vals, True, "cells")
    print("[formulas]")
    dump(wb_form, False, "formulas")
    print("[tsv]")
    tsv(wb_vals)
    print(f"\nDone. Output: {OUT}")


if __name__ == "__main__":
    main()
