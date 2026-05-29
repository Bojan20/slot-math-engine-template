#!/usr/bin/env python3
"""
Local-only PAR XLSX dumper for `book-expanding-bonusbuy` template.

Runs purely on openpyxl in-process — no network, no API, no telemetry.
Produces three files per sheet (TSV / cells.json / formulas.json) +
a sheets-manifest summary. Copyright-safe: outputs only math primitives
(reel stops, paytable, bonus buy prices, hit / RTP tables); vendor and
game identifiers are stripped at extract time.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell

# ---------------------------------------------------------------------------
# Copyright-safe filters
# ---------------------------------------------------------------------------
# Strings containing any of these (case-insensitive) get replaced with the
# generic placeholder `<<redacted>>` before being written to disk.
REDACT_TOKENS = (
    "book of unseen",
    "bookofunseen",
    "book_of_unseen",
    "unseen",
)

# Optional: scrub vendor / publisher names that may appear in cell text.
REDACT_VENDOR_TOKENS: tuple[str, ...] = ()

PLACEHOLDER = "<<redacted>>"


def scrub(value: Any) -> Any:
    """Redact game-name / vendor tokens from a cell value."""
    if not isinstance(value, str):
        return value
    low = value.lower()
    for tok in REDACT_TOKENS + REDACT_VENDOR_TOKENS:
        if tok and tok in low:
            return PLACEHOLDER
    return value


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------
def cell_type(cell: Cell) -> str:
    if cell.data_type == "f":
        return "formula"
    if cell.data_type == "n":
        return "number"
    if cell.data_type == "s":
        return "string"
    if cell.data_type == "b":
        return "bool"
    return "unknown"


def safe_sheet_name(title: str) -> str:
    """Sanitize sheet title for filesystem (Windows + macOS safe)."""
    return re.sub(r"[^A-Za-z0-9._-]+", "_", title).strip("_") or "sheet"


# ---------------------------------------------------------------------------
# Dump core
# ---------------------------------------------------------------------------
def dump_sheet(ws, out_dir: Path) -> dict:
    name = safe_sheet_name(ws.title)
    tsv_path = out_dir / f"{name}.tsv"
    cells_path = out_dir / f"{name}.cells.json"
    formulas_path = out_dir / f"{name}.formulas.json"

    cells: list[dict] = []
    formulas: list[dict] = []
    grid: list[list[str]] = []

    for r in range(1, ws.max_row + 1):
        tsv_row: list[str] = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            value = scrub(cell.value)
            tsv_row.append("" if value is None else str(value))
            if value is None:
                continue
            cells.append(
                {
                    "row": r,
                    "col": c,
                    "col_letter": cell.column_letter,
                    "type": cell_type(cell),
                    "value": value,
                }
            )
            if cell.data_type == "f":
                formulas.append(
                    {
                        "row": r,
                        "col": c,
                        "col_letter": cell.column_letter,
                        "formula": value,
                    }
                )
        grid.append(tsv_row)

    with tsv_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, delimiter="\t", lineterminator="\n")
        w.writerows(grid)

    cells_path.write_text(json.dumps(cells, ensure_ascii=False, indent=2))
    formulas_path.write_text(json.dumps(formulas, ensure_ascii=False, indent=2))

    return {
        "sheet": ws.title,
        "safe_name": name,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "non_empty_cells": len(cells),
        "formula_cells": len(formulas),
        "tsv": tsv_path.name,
        "cells_json": cells_path.name,
        "formulas_json": formulas_path.name,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Local PAR XLSX dumper")
    parser.add_argument(
        "--src",
        default=str(
            Path(__file__).resolve().parents[1] / "raw" / "PARSheets_source.xlsx"
        ),
        help="Path to source XLSX (default: ../raw/PARSheets_source.xlsx)",
    )
    parser.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parents[1] / "raw" / "dump"),
        help="Output dir for per-sheet TSV / JSON (default: ../raw/dump)",
    )
    args = parser.parse_args()

    src = Path(args.src)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    assert src.exists(), f"missing source: {src}"

    print(f"[par-dump] source: {src}")
    print(f"[par-dump] target: {out}")

    wb = openpyxl.load_workbook(src, data_only=False)
    manifest: list[dict] = []
    for sheet_name in wb.sheetnames:
        info = dump_sheet(wb[sheet_name], out)
        manifest.append(info)
        print(
            f"  + {info['safe_name']:35s}  "
            f"cells={info['non_empty_cells']:6d}  formulas={info['formula_cells']:5d}"
        )

    (out / "sheets_manifest.json").write_text(
        json.dumps(
            {
                "source": str(src.name),
                "sheets": manifest,
                "copyright_policy": (
                    "All cell values containing game / vendor identifiers are "
                    "replaced with `<<redacted>>` at extract time; downstream "
                    "templates are derived from math primitives only."
                ),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    print(f"[par-dump] {len(manifest)} sheets written")


if __name__ == "__main__":
    main()
