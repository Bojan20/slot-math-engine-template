#!/usr/bin/env python3
"""
Semantic extractor — izvlaci strukturirane entitete iz PAR Excel fajlova.
Generise agent-friendly JSON corpus.
"""
import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).parent.parent
OUT_DIR = REPO / "agents" / "math-agent" / "corpus"

RAW_PATHS = {
    "skeleton-key": REPO / "games" / "skeleton-key" / "raw" / "PARSheets_SkeletonKey.xlsx",
    "fortune-coin-boost-classic": REPO / "games" / "fortune-coin-boost-classic" / "raw" / "ParSheets_FortuneCoinBoost_Classic.xlsx",
}


def clean_val(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "--" or v == "-":
            return None
        return v
    return v


def extract_metadata(ws) -> dict:
    """Metadata je u prvih 5 redova, kolone A-O."""
    meta = {}
    for r in range(1, 6):
        for c in range(1, 20):
            v = clean_val(ws.cell(row=r, column=c).value)
            if v is None:
                continue
            s = str(v)
            if "Software ID:" in s:
                # iduci cell je vrednost
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt:
                    meta["swid"] = str(nxt)
            if "Hold" in s and ":" in s:
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    meta["hold_pct"] = float(nxt) if isinstance(nxt, (int, float)) else None
            if "Hit Frequency" in s and ":" in s:
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    meta["hit_frequency"] = float(nxt) if isinstance(nxt, (int, float)) else None
            if "Win Frequency" in s and ":" in s:
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    meta["win_frequency"] = float(nxt) if isinstance(nxt, (int, float)) else None
            if "All Ways Win Frequency" in s and ":" in s:
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    meta["all_ways_win_frequency"] = float(nxt) if isinstance(nxt, (int, float)) else None
            if "All Ways Hit Frequency" in s and ":" in s:
                nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    meta["all_ways_hit_frequency"] = float(nxt) if isinstance(nxt, (int, float)) else None
    return meta


def find_paytable_bounds(ws, start_row=1) -> tuple:
    """Nadji pocetak i kraj paytable sekcije."""
    header_row = None
    end_row = None
    for r in range(start_row, min(ws.max_row + 1, start_row + 200)):
        row_text = " ".join(
            str(clean_val(ws.cell(row=r, column=c).value) or "")
            for c in range(1, min(ws.max_column + 1, 20))
        ).lower()
        if "combination" in row_text and ("pay" in row_text or "pays" in row_text):
            header_row = r
            continue
        if header_row and end_row is None:
            # Kraj je kada naidjemo na prazan red ili novu sekciju
            all_empty = True
            has_content = False
            for c in range(1, min(ws.max_column + 1, 15)):
                v = clean_val(ws.cell(row=r, column=c).value)
                if v is not None:
                    all_empty = False
                    has_content = True
                    break
            if all_empty and has_content is False:
                # Proveri da li je sledeci red isto prazan — kraj paytable
                nxt_empty = True
                for c in range(1, min(ws.max_column + 1, 15)):
                    if clean_val(ws.cell(row=r + 1, column=c).value) is not None:
                        nxt_empty = False
                        break
                if nxt_empty:
                    end_row = r - 1
                    break
            # Ako naidjemo na "Scatter" u koloni C (ili E), to je jos uvek paytable
            # Ako naidjemo na "Base Game multiway RTP" — kraj
            if "rtp" in row_text and "base" in row_text:
                end_row = r - 1
                break
            if "free spins" in row_text and "bonus" in row_text:
                end_row = r - 1
                break
    if header_row and end_row is None:
        end_row = header_row + 100
    return header_row, end_row


def extract_paytable(ws, header_row: int, end_row: int) -> list:
    """Ekstraktuj paytable redove."""
    rows = []
    # Nadji header kolone
    headers = {}
    for c in range(1, min(ws.max_column + 1, 20)):
        h = clean_val(ws.cell(row=header_row, column=c).value)
        if h:
            headers[c] = str(h)
    for r in range(header_row + 1, end_row + 1):
        row_data = {"_row": r}
        has_val = False
        for c, h in headers.items():
            v = clean_val(ws.cell(row=r, column=c).value)
            if v is not None:
                has_val = True
                # Konvertuj brojeve
                if isinstance(v, (int, float)):
                    row_data[h] = v
                else:
                    row_data[h] = v
        if has_val:
            rows.append(row_data)
    return rows


def extract_rtp_breakdown(ws) -> dict:
    """Nadji RTP breakdown redove (sadrze 'RTP' ili 'Return %')."""
    breakdown = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 25)):
            v = clean_val(ws.cell(row=r, column=c).value)
            if v and isinstance(v, str) and "rtp" in v.lower():
                nxt = clean_val(ws.cell(row=r, column=c + 2).value)
                if nxt is None:
                    nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                if nxt is not None:
                    try:
                        breakdown[v.strip().rstrip(":")] = float(nxt)
                    except (ValueError, TypeError):
                        breakdown[v.strip().rstrip(":")] = nxt
    return breakdown


def extract_reel_strips(ws) -> dict:
    """Nadji reel strip sekcije — red sa 'Reel 1', 'Reel 2', ..."""
    strips = {}
    in_strip = False
    start_row = None
    reel_cols = {}
    for r in range(1, min(ws.max_row + 1, 1600)):
        row_vals = []
        for c in range(1, min(ws.max_column + 1, 120)):
            v = clean_val(ws.cell(row=r, column=c).value)
            if v is not None:
                row_vals.append((c, v))
        # Detekcija headera
        for c, v in row_vals:
            if isinstance(v, str) and re.match(r'^Reel\s*\d+$', v.strip()):
                reel_cols[c] = v.strip()
                start_row = r
                in_strip = True
        if in_strip and start_row and r > start_row:
            # Kraj stripa: red gde su sve kolone prazne
            has_any = False
            for c in reel_cols:
                v = clean_val(ws.cell(row=r, column=c).value)
                if v is not None:
                    has_any = True
                    if reel_cols[c] not in strips:
                        strips[reel_cols[c]] = []
                    strips[reel_cols[c]].append(v)
            if not has_any:
                # Prazan red = kraj stripa
                in_strip = False
                reel_cols = {}
    return strips


def extract_bonus_info(ws) -> dict:
    """Izdvoji Free Spins / Bonus podatke."""
    bonus = {}
    in_fs = False
    for r in range(1, min(ws.max_row + 1, 200)):
        row_text = " ".join(
            str(clean_val(ws.cell(row=r, column=c).value) or "")
            for c in range(1, 20)
        ).lower()
        if "free spins" in row_text and "bonus" in row_text:
            in_fs = True
        if in_fs:
            # Trazi trigger info
            for c in range(1, 20):
                v = clean_val(ws.cell(row=r, column=c).value)
                if v and isinstance(v, str) and "free spins" in v.lower():
                    nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                    if nxt is not None:
                        bonus[v.strip()] = float(nxt) if isinstance(nxt, (int, float)) else nxt
            # Avg # Free Spins
            for c in range(1, 20):
                v = clean_val(ws.cell(row=r, column=c).value)
                if v and isinstance(v, str) and "ave # free spins" in v.lower():
                    nxt = clean_val(ws.cell(row=r, column=c + 1).value)
                    if nxt is not None:
                        bonus["avg_free_spins"] = float(nxt) if isinstance(nxt, (int, float)) else nxt
            # Notes
            if "notes:" in row_text:
                notes = []
                for nr in range(r, min(r + 20, ws.max_row + 1)):
                    note_text = clean_val(ws.cell(row=nr, column=3).value)
                    if note_text:
                        notes.append(str(note_text))
                bonus["notes"] = notes
    return bonus


def process_sheet(ws, sheet_name: str) -> dict:
    """Procesiraj jedan worksheet."""
    print(f"    [semantic] {sheet_name}")
    meta = extract_metadata(ws)
    header_row, end_row = find_paytable_bounds(ws)
    paytable = []
    if header_row and end_row:
        paytable = extract_paytable(ws, header_row, end_row)
        print(f"      paytable: R{header_row}-R{end_row} ({len(paytable)} rows)")
    rtp = extract_rtp_breakdown(ws)
    strips = extract_reel_strips(ws)
    if strips:
        print(f"      reel strips: {list(strips.keys())} ({len(list(strips.values())[0]) if strips else 0} positions each)")
    bonus = extract_bonus_info(ws)
    return {
        "sheet_name": sheet_name,
        "metadata": meta,
        "paytable": paytable,
        "rtp_breakdown": rtp,
        "reel_strips": strips,
        "bonus": bonus,
    }


def process_game(game_key: str, path: Path) -> dict:
    print(f"[PROCESS] {game_key}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        sheet_data = process_sheet(ws, ws.title)
        sheets.append(sheet_data)
    return {
        "game_key": game_key,
        "source": path.name,
        "sheets": sheets,
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for game_key, path in RAW_PATHS.items():
        if not path.exists():
            print(f"[SKIP] {path}")
            continue
        data = process_game(game_key, path)
        game_dir = OUT_DIR / game_key
        game_dir.mkdir(exist_ok=True)

        # Upisi strukturirane fajlove
        for idx, sheet in enumerate(data["sheets"]):
            suffix = sheet["sheet_name"].replace(" ", "_").replace("/", "_")
            with open(game_dir / f"sheet_{idx}_{suffix}.json", "w", encoding="utf-8") as f:
                json.dump(sheet, f, indent=2, ensure_ascii=False, default=str)

        # Generisi summary
        summary = {
            "game_key": game_key,
            "swids": [s["metadata"].get("swid") for s in data["sheets"] if s["metadata"].get("swid")],
            "hold_range": [s["metadata"].get("hold_pct") for s in data["sheets"] if s["metadata"].get("hold_pct")],
            "hit_freq_range": [s["metadata"].get("hit_frequency") for s in data["sheets"] if s["metadata"].get("hit_frequency")],
            "sheet_count": len(data["sheets"]),
            "total_paytable_rows": sum(len(s["paytable"]) for s in data["sheets"]),
            "total_reel_positions": sum(
                len(v) for s in data["sheets"] for v in s["reel_strips"].values()
            ),
        }
        with open(game_dir / "summary.json", "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2, ensure_ascii=False, default=str)

        print(f"[WRITE] {game_dir}")
    print("[DONE]")


if __name__ == "__main__":
    main()
