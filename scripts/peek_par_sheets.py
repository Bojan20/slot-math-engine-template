#!/usr/bin/env python3
"""Peek prvih N redova svakog PAR sheet-a da vidimo strukturu."""
import openpyxl
from pathlib import Path

REPO = Path(__file__).parent.parent
paths = {
    "skeleton-key": REPO / "games/skeleton-key/raw/PARSheets_SkeletonKey.xlsx",
    "fortune-coin": REPO / "games/fortune-coin-boost-classic/raw/ParSheets_FortuneCoinBoost_Classic.xlsx",
}

for game, p in paths.items():
    print(f"\n{'='*60}\n{game.upper()}\n{'='*60}")
    wb = openpyxl.load_workbook(p, data_only=True)
    for ws in wb.worksheets:
        print(f"\n--- Sheet: '{ws.title}' ({ws.max_row}r x {ws.max_column}c) ---")
        for r in range(1, min(31, ws.max_row + 1)):
            row_vals = []
            for c in range(1, min(21, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_vals.append(f"{openpyxl.utils.get_column_letter(c)}={v}")
            if row_vals:
                print(f"  R{r}: {' | '.join(row_vals[:10])}")
        print("  ...")
