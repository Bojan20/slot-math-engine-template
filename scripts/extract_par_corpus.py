#!/usr/bin/env python3
"""
Ultimativna ekstrakcija svake ćelije iz PAR Excel fajlova.
Generiše strukturirani JSON corpus za lokalni math-agent.
"""
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell

REPO_ROOT = Path(__file__).parent.parent
OUT_DIR = REPO_ROOT / "agents" / "math-agent" / "corpus"
RAW_PATHS = {
    "skeleton-key": REPO_ROOT / "games" / "skeleton-key" / "raw" / "PARSheets_SkeletonKey.xlsx",
    "fortune-coin-boost-classic": REPO_ROOT / "games" / "fortune-coin-boost-classic" / "raw" / "ParSheets_FortuneCoinBoost_Classic.xlsx",
}


def cell_value(cell: Cell) -> Any:
    """Vrati čistu vrednost ćelije."""
    if cell.value is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value


def cell_type(cell: Cell) -> str:
    if cell.data_type == "f":
        return "formula"
    if cell.data_type == "n":
        return "number"
    if cell.data_type == "s":
        return "string"
    if cell.data_type == "b":
        return "bool"
    return "unknown"


def extract_sheet(ws) -> dict:
    """Izvuči apsolutno sve iz jednog worksheet-a."""
    rows = []
    max_row = ws.max_row
    max_col = ws.max_column
    # Iteriraj svaku ćeliju
    for r in range(1, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell_value(cell)
            if val is not None:
                row_cells.append({
                    "col": c,
                    "col_letter": cell.column_letter,
                    "value": val,
                    "type": cell_type(cell),
                    "formula": cell.value if cell.data_type == "f" else None,
                })
        if row_cells:
            rows.append({"row": r, "cells": row_cells})
    return {
        "title": ws.title,
        "max_row": max_row,
        "max_col": max_col,
        "rows": rows,
    }


def classify_sheet(title: str, rows: list) -> dict:
    """Heuristička klasifikacija sheet-a na osnovu sadržaja."""
    title_lower = title.lower()
    hints = []
    if "par" in title_lower:
        hints.append("par_sheet")
    if "base" in title_lower:
        hints.append("base_game")
    if "bonus" in title_lower:
        hints.append("bonus")
    if "free" in title_lower or "fs" in title_lower:
        hints.append("free_spins")
    if "pay" in title_lower:
        hints.append("paytable")
    if "reel" in title_lower:
        hints.append("reel_strip")
    if "weight" in title_lower:
        hints.append("weights")
    # Pogledaj prve redove za dodatne hintove
    all_text = " ".join(
        str(c["value"]).lower()
        for r in rows[:30]
        for c in r["cells"]
        if c["value"] is not None
    )
    if "rtp" in all_text:
        hints.append("has_rtp")
    if "hit frequency" in all_text or "hit freq" in all_text:
        hints.append("has_hit_freq")
    if "symbol" in all_text and "count" in all_text:
        hints.append("paytable")
    if "reel" in all_text and "strip" in all_text:
        hints.append("reel_strip")
    return {"hints": list(set(hints)), "preview_text": all_text[:500]}


def extract_game(game_key: str, raw_path: Path) -> dict:
    print(f"[EXTRACT] {game_key} ← {raw_path}")
    wb = openpyxl.load_workbook(raw_path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        sheet_data = extract_sheet(ws)
        classification = classify_sheet(ws.title, sheet_data["rows"])
        sheets.append({
            "sheet_name": ws.title,
            "dimensions": f"{ws.dimensions}",
            "classification": classification,
            "data": sheet_data,
        })
        print(f"  ✓ Sheet '{ws.title}': {ws.max_row} rows × {ws.max_column} cols")
    return {
        "game_key": game_key,
        "source_filename": raw_path.name,
        "source_path": str(raw_path.relative_to(REPO_ROOT)),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for game_key, raw_path in RAW_PATHS.items():
        if not raw_path.exists():
            print(f"[SKIP] Missing {raw_path}")
            continue
        corpus = extract_game(game_key, raw_path)
        out_file = OUT_DIR / game_key / "full_corpus.json"
        out_file.parent.mkdir(parents=True, exist_ok=True)
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(corpus, f, indent=2, ensure_ascii=False, default=str)
        print(f"[WRITE] {out_file} ({out_file.stat().st_size / 1024 / 1024:.2f} MB)")
    print("[DONE] Sva izvučeno.")


if __name__ == "__main__":
    main()
