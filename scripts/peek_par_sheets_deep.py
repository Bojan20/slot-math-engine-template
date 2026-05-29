#!/usr/bin/env python3
"""Deep peek — key sections."""
import openpyxl
from pathlib import Path

REPO = Path(__file__).parent.parent
p = REPO / "games/fortune-coin-boost-classic/raw/ParSheets_FortuneCoinBoost_Classic.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['par_001']

print("=== Rows 1-60 ===")
for r in range(1, 61):
    vals = []
    for c in range(1, 15):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals.append(f"{openpyxl.utils.get_column_letter(c)}={v}")
    if vals:
        print(f"R{r}: {' | '.join(vals)}")

print("\n=== Rows 100-160 ===")
for r in range(100, 161):
    vals = []
    for c in range(1, 15):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals.append(f"{openpyxl.utils.get_column_letter(c)}={v}")
    if vals:
        print(f"R{r}: {' | '.join(vals)}")

print("\n=== Rows 300-360 ===")
for r in range(300, 361):
    vals = []
    for c in range(1, 15):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals.append(f"{openpyxl.utils.get_column_letter(c)}={v}")
    if vals:
        print(f"R{r}: {' | '.join(vals)}")

# Find reel strip section
print("\n=== Searching for 'Reel' keyword ===")
for r in range(1, ws.max_row + 1):
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, str) and "Reel" in v and "Combinations" not in v:
            print(f"R{r} C{c}: {v}")
            if r > 500:
                break
    if r > 500:
        break
