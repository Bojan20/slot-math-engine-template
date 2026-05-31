"""Generic XLSX PAR adapter — Faza 1.4.

Parses vendor-agnostic XLSX layouts that follow common slot PAR conventions:
  * Reel strips sheet  → reels.{reel_id}.strip[]
  * Paytable sheet     → paytable.{symbol}.{count}
  * Summary sheet      → meta.{rtp_target_pct, volatility, max_win_x_bet}

Heuristic-driven; no vendor-specific magic numbers. For vendor-specific adapters
(e.g. IGT, Pragmatic) extend this or register a dedicated adapter.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from tools.par_normalize.adapters import register

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise RuntimeError("openpyxl is required for XLSX adapter") from exc


# Sheet name heuristics (case-insensitive substring match)
_REEL_SHEET_HINTS = ("reel", "strip", "strips", "symbols")
_PAYTABLE_SHEET_HINTS = ("pay", "pays", "paytable", "table", "awards")
_SUMMARY_SHEET_HINTS = ("summary", "stat", "rtp", "overview", "info", "game")

# Reel column header hints
_REEL_HEADER_RE = re.compile(r"reel\s*(\d+)", re.IGNORECASE)

# Numeric extraction helpers
_RTP_RE = re.compile(r"(\d{2}\.\d{1,4})")
_VOLATILITY_RE = re.compile(r"(very[ _-]?low|low|med[ _-]?low|med|med[ _-]?high|high|very[ _-]?high|extreme)", re.IGNORECASE)
_MAX_WIN_RE = re.compile(r"max[ _-]?win.*?([\d,]+)", re.IGNORECASE)


def _score_sheet(name: str, hints: Tuple[str, ...]) -> int:
    lowered = name.lower()
    return sum(1 for h in hints if h in lowered)


def _best_sheet(wb: openpyxl.Workbook, hints: Tuple[str, ...]) -> Optional[str]:
    scores = [(s, _score_sheet(s, hints)) for s in wb.sheetnames]
    scores.sort(key=lambda x: x[1], reverse=True)
    return scores[0][0] if scores and scores[0][1] > 0 else None


def _cell_value(cell: Any) -> Any:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _is_numeric_row(row: List[Any]) -> bool:
    """True if majority of non-None cells are numeric."""
    vals = [v for v in row if v is not None]
    if not vals:
        return False
    numeric = sum(1 for v in vals if isinstance(v, (int, float)))
    return numeric >= len(vals) // 2


def _parse_reel_strips(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, List[str]]:
    """Extract reel strips from a worksheet.

    Two supported layouts:
      A) Header row with 'Reel 1', 'Reel 2', ... → columns are reels.
      B) No header → first N columns are reels, rows are positions.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Try layout A: detect header row
    header_row_idx = None
    reel_cols: Dict[str, int] = {}
    for ridx, row in enumerate(rows[:5]):
        for cidx, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip()
            m = _REEL_HEADER_RE.match(text)
            if m:
                reel_id = m.group(1)
                reel_cols[reel_id] = cidx
        if reel_cols:
            header_row_idx = ridx
            break

    if reel_cols:
        data_start = header_row_idx + 1
        reels: Dict[str, List[str]] = {rid: [] for rid in reel_cols}
        for row in rows[data_start:]:
            for rid, cidx in reel_cols.items():
                val = row[cidx] if cidx < len(row) else None
                if val is not None:
                    reels[rid].append(str(val).strip())
        return reels

    # Layout B: assume first non-empty row is data, each column is a reel
    # Skip purely numeric rows (might be stats)
    for ridx, row in enumerate(rows):
        if any(v is not None and not isinstance(v, (int, float)) for v in row):
            # Use all columns that have at least one string/symbol
            reels: Dict[str, List[str]] = {}
            for cidx in range(len(row)):
                col_vals = []
                for rr in rows[ridx:]:
                    if cidx < len(rr) and rr[cidx] is not None:
                        col_vals.append(str(rr[cidx]).strip())
                if col_vals:
                    reels[str(cidx + 1)] = col_vals
            return reels

    return {}


def _parse_paytable(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Dict[str, Any]]:
    """Extract paytable: symbol → {3: x, 4: y, 5: z, ...}.

    Expects first row to contain headers like 'Symbol', '3', '4', '5' or
    '3 of a kind', '4 of a kind', etc.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Find header row
    header_row_idx = None
    count_cols: Dict[int, int] = {}  # count -> column index
    symbol_col: Optional[int] = None

    for ridx, row in enumerate(rows[:5]):
        for cidx, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip().lower()
            if text in ("symbol", "sym", "icon", "name"):
                symbol_col = cidx
            # Extract count numbers from header
            m = re.search(r"^(\d+)", text)
            if m and "kind" in text or text.isdigit():
                count_cols[int(m.group(1))] = cidx
            elif text.isdigit():
                count_cols[int(text)] = cidx
        if symbol_col is not None and count_cols:
            header_row_idx = ridx
            break

    # Fallback: if no symbol column but numeric headers, assume col 0 is symbol
    if symbol_col is None and count_cols:
        symbol_col = 0
        header_row_idx = header_row_idx or 0

    if header_row_idx is None:
        return {}

    paytable: Dict[str, Dict[str, Any]] = {}
    for row in rows[header_row_idx + 1:]:
        if symbol_col is None or symbol_col >= len(row) or row[symbol_col] is None:
            continue
        sym = str(row[symbol_col]).strip()
        if not sym:
            continue
        entry: Dict[str, Any] = {}
        for count, cidx in count_cols.items():
            if cidx < len(row) and row[cidx] is not None:
                try:
                    entry[str(count)] = float(row[cidx])
                except (ValueError, TypeError):
                    entry[str(count)] = row[cidx]
        if entry:
            paytable[sym] = entry

    return paytable


def _parse_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
    """Extract meta fields from summary-like sheet.

    Scans all cells for RTP, volatility, max win, game name.
    Also looks for key-value pairs in first two columns.
    """
    meta: Dict[str, Any] = {}
    game_name_candidates: List[str] = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Key-value pattern in first two columns
        if len(row) >= 2 and row[0] is not None and row[1] is not None:
            key = str(row[0]).strip().lower().replace(" ", "_")
            val = row[1]
            if "rtp" in key and isinstance(val, (int, float)):
                meta["rtp_target_pct"] = float(val) if float(val) < 100 else float(val) / 100.0
            elif "volatil" in key and isinstance(val, str):
                meta["volatility"] = val.lower().replace(" ", "_")
            elif "max_win" in key and isinstance(val, (int, float)):
                meta["max_win_x_bet"] = int(val)
            elif "game" in key and isinstance(val, str):
                game_name_candidates.append(val.strip())

        # Regex scan across all cells in row
        for cell in row:
            if cell is None:
                continue
            text = str(cell)
            if "rtp" in text.lower():
                m = _RTP_RE.search(text)
                if m and "rtp_target_pct" not in meta:
                    v = float(m.group(1))
                    meta["rtp_target_pct"] = v if v < 100 else v / 100.0
            if "volatil" in text.lower():
                m = _VOLATILITY_RE.search(text)
                if m and "volatility" not in meta:
                    meta["volatility"] = m.group(1).lower().replace(" ", "_").replace("-", "_")
            if "max" in text.lower() and "win" in text.lower():
                m = _MAX_WIN_RE.search(text)
                if m and "max_win_x_bet" not in meta:
                    meta["max_win_x_bet"] = int(m.group(1).replace(",", ""))
            if len(text) < 60 and text.strip() and not any(k in text.lower() for k in ("rtp", "volatil", "max win", "hit freq", "return")):
                # Could be game name if it's a short prominent string in early rows
                pass

    if game_name_candidates:
        meta["game_name"] = game_name_candidates[0]

    return meta


def adapt(path: Path | str) -> dict:
    """Parse generic XLSX PAR sheet into canonical dict."""
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)

    # Select sheets heuristically
    reel_sheet_name = _best_sheet(wb, _REEL_SHEET_HINTS)
    paytable_sheet_name = _best_sheet(wb, _PAYTABLE_SHEET_HINTS)
    summary_sheet_name = _best_sheet(wb, _SUMMARY_SHEET_HINTS)

    reels: Dict[str, List[str]] = {}
    if reel_sheet_name:
        reels = _parse_reel_strips(wb[reel_sheet_name])

    paytable: Dict[str, Dict[str, Any]] = {}
    if paytable_sheet_name:
        paytable = _parse_paytable(wb[paytable_sheet_name])

    meta: Dict[str, Any] = {}
    if summary_sheet_name:
        meta = _parse_summary(wb[summary_sheet_name])

    wb.close()

    # Build canonical structure
    canonical: Dict[str, Any] = {
        "schema": "slot-math-canonical-par/v1",
        "meta": {
            "game_name": meta.get("game_name", p.stem),
            "variant_id": "default",
            "rtp_target_pct": meta.get("rtp_target_pct", 96.0),
        },
        "topology": {
            "type": "lines",
            "reel_count": len(reels),
            "rows_per_reel": max(len(strip) for strip in reels.values()) if reels else 3,
            "paylines": 10,
        },
        "reels": {
            rid: {"strip": strip}
            for rid, strip in reels.items()
        },
        "paytable": paytable,
        "rtp": {
            "target_pct": meta.get("rtp_target_pct", 96.0),
            "base_pct": None,
            "feature_pct": None,
            "hit_frequency_pct": None,
        },
        "rng_profile": {
            "algorithm": "Pcg64",
            "jurisdiction": "MGA",
        },
        "source": {
            "vendor": "generic",
            "adapter_version": "generic_xlsx/1.0.0",
        },
    }

    if "volatility" in meta:
        canonical["meta"]["volatility"] = meta["volatility"]
    if "max_win_x_bet" in meta:
        canonical["meta"]["max_win_x_bet"] = meta["max_win_x_bet"]

    # Add derived fields if paytable present
    if paytable:
        canonical["symbols"] = list(paytable.keys())

    return canonical


register("xlsx", adapt)
