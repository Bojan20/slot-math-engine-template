"""Exhaustive XLSX → JSON extractor.

Two-pass parse:
  Pass 1: data_only=False — captures formula text in every formula cell.
  Pass 2: data_only=True  — captures the cached evaluated value.

Per sheet we emit:
  <sheet>.cells.json    — every non-empty cell with full attribute matrix
  <sheet>.layout.json   — merged ranges, frozen panes, hidden rows/cols,
                          column widths, row heights, print areas, autofilter
  <sheet>.styles.json   — per-cell font/fill/border/alignment + conditional
                          formats; deduplicated style table for compactness
  <sheet>.tables.json   — Excel tables, autofilters
  <sheet>.charts.json   — chart kinds + series + ranges
  <sheet>.validation.json — data validations (dropdowns, ranges)

Workbook-level:
  workbook.json — defined names, named ranges, props (creator, modified),
                  sheet order, security, custom doc props

Plus aggregate:
  ultimate_extract.summary.json — sheet counts, cell counts, style counts
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


# ─── Lenient loader — tolerates IGT/vendor sheets that violate openpyxl's
# strict validation (e.g. text_rotation > 180, extended border styles, etc.)
def _patch_openpyxl_for_vendor_sheets() -> None:
    """Soften openpyxl validators that reject otherwise-valid vendor XLSX.

    Some vendors emit Excel files via custom tools that produce values
    outside openpyxl's strict bounds (e.g. text_rotation=255 to encode
    'vertical stack' instead of 0..180). Calling this once before load
    makes the descriptors clamp instead of raise.
    """
    from openpyxl.descriptors.base import MinMax, Set

    # text_rotation: clamp to 0..180.
    _orig_minmax_set = MinMax.__set__

    def _lenient_minmax_set(self, instance, value):
        try:
            _orig_minmax_set(self, instance, value)
        except (ValueError, TypeError):
            try:
                if value is None:
                    object.__setattr__(instance, self.name, None)
                    return
                v = float(value)
                if hasattr(self, "min") and self.min is not None and v < self.min:
                    v = self.min
                if hasattr(self, "max") and self.max is not None and v > self.max:
                    v = self.max
                if hasattr(self, "expected_type") and self.expected_type is int:
                    v = int(v)
                # Bypass setter again — write directly.
                object.__setattr__(instance, self.name, v)
            except Exception:
                # Last resort: skip silently.
                pass

    MinMax.__set__ = _lenient_minmax_set

    # Set validator: if value is not in the allowed set, fall back to default
    # (first allowed value) instead of raising.
    _orig_set_set = Set.__set__

    def _lenient_set_set(self, instance, value):
        try:
            _orig_set_set(self, instance, value)
        except (ValueError, TypeError):
            try:
                fallback = next(iter(self.values))
                object.__setattr__(instance, self.name, fallback)
            except Exception:
                pass

    Set.__set__ = _lenient_set_set


# Idempotent — call once at import time.
_patch_openpyxl_for_vendor_sheets()


# ─── Dataclasses for stats ────────────────────────────────────────────────


@dataclass
class SheetStats:
    name: str
    max_row: int
    max_col: int
    cells_total: int = 0
    cells_with_value: int = 0
    cells_with_formula: int = 0
    cells_with_comment: int = 0
    cells_with_hyperlink: int = 0
    merged_count: int = 0
    hidden_rows: int = 0
    hidden_cols: int = 0
    style_records: int = 0
    table_count: int = 0
    chart_count: int = 0
    validation_count: int = 0


@dataclass
class ExtractionStats:
    workbook: str
    sheet_count: int
    defined_names: int
    total_cells: int
    total_formulas: int
    total_comments: int
    total_hyperlinks: int
    out_dir: str
    sheets: list[SheetStats] = field(default_factory=list)


# ─── Cell serializer ──────────────────────────────────────────────────────


def _safe_serialise(value: Any) -> Any:
    """Convert openpyxl cell values to JSON-safe equivalents.

    Handles datetime, Decimal, bytes, named tuple, and falls back to str().
    """
    import datetime
    import decimal

    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        # NaN / inf are valid JSON floats only in some parsers; coerce to string.
        if isinstance(value, float):
            if value != value:  # NaN
                return {"__nan__": True}
            if value == float("inf"):
                return {"__inf__": "+"}
            if value == float("-inf"):
                return {"__inf__": "-"}
        return value
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return {"__date__": value.isoformat()}
    if isinstance(value, decimal.Decimal):
        return {"__decimal__": str(value)}
    if isinstance(value, bytes):
        try:
            return {"__bytes__": value.decode("utf-8", errors="replace")}
        except Exception:
            return {"__bytes_hex__": value.hex()}
    # Fallback for any exotic openpyxl types (ArrayFormula, etc).
    try:
        return {"__repr__": str(value), "__type__": type(value).__name__}
    except Exception:
        return {"__unrepresentable__": True, "__type__": type(value).__name__}


def _color_to_str(color) -> Any:
    """Color → JSON-serialisable string ("FFAABB22" or "theme:N+tint:0.5" etc)."""
    if color is None:
        return None
    # openpyxl Color has type-discriminated fields: rgb, theme, indexed.
    try:
        # Some openpyxl versions wrap rgb in an RGB(...) named tuple.
        rgb = getattr(color, "rgb", None)
        if rgb is not None:
            return str(rgb)
        if getattr(color, "type", None) == "theme":
            t = getattr(color, "theme", None)
            tint = getattr(color, "tint", 0.0)
            return f"theme:{t}+tint:{tint}"
        if getattr(color, "type", None) == "indexed":
            return f"indexed:{getattr(color, 'indexed', None)}"
    except Exception:
        pass
    return str(color)


def _font_to_dict(font) -> dict:
    try:
        return {
            "name": font.name,
            "size": float(font.size) if font.size is not None else None,
            "bold": bool(font.bold) if font.bold is not None else None,
            "italic": bool(font.italic) if font.italic is not None else None,
            "underline": font.underline,
            "strike": bool(font.strike) if font.strike is not None else None,
            "color": _color_to_str(font.color) if font.color else None,
        }
    except Exception:
        return {}


def _fill_to_dict(fill) -> dict:
    try:
        return {
            "patternType": fill.patternType,
            "fgColor": _color_to_str(fill.fgColor) if fill.fgColor else None,
            "bgColor": _color_to_str(fill.bgColor) if fill.bgColor else None,
        }
    except Exception:
        return {}


def _border_to_dict(border) -> dict:
    try:
        out = {}
        for side_name in ("left", "right", "top", "bottom", "diagonal"):
            side = getattr(border, side_name, None)
            if side is not None:
                out[side_name] = {
                    "style": side.style,
                    "color": _color_to_str(side.color) if side.color else None,
                }
        return out
    except Exception:
        return {}


def _alignment_to_dict(al) -> dict:
    try:
        return {
            "horizontal": al.horizontal,
            "vertical": al.vertical,
            "wrap_text": al.wrap_text,
            "text_rotation": al.text_rotation,
            "indent": al.indent,
        }
    except Exception:
        return {}


def _style_signature(font_d, fill_d, border_d, align_d, number_format) -> str:
    """Stable hashable signature for style deduplication."""
    return json.dumps(
        {"f": font_d, "fl": fill_d, "b": border_d, "a": align_d, "n": number_format},
        sort_keys=True,
        default=str,
    )


# ─── Per-sheet extractor ──────────────────────────────────────────────────


def _extract_sheet_cells(
    ws_formula: Worksheet,
    ws_value: Worksheet,
    sheet_dir: Path,
    stats: SheetStats,
) -> None:
    """Walk every cell in the sheet (formula pass + value pass) and emit JSON.

    `ws_formula` is from the data_only=False workbook (has formula text).
    `ws_value`   is from the data_only=True  workbook (has evaluated values).
    """
    cells: dict[str, dict] = {}
    style_table: dict[str, int] = {}  # signature → id
    cell_styles: dict[str, int] = {}

    max_row = ws_formula.max_row or 0
    max_col = ws_formula.max_column or 0

    # Build (row, col) → value-cell lookup if the sheet exists in value pass.
    # ws_value may differ in dimensions if there are formulas-only cells.
    value_lookup: dict[tuple[int, int], Any] = {}
    if ws_value is not None:
        v_max_row = ws_value.max_row or 0
        v_max_col = ws_value.max_column or 0
        for r in range(1, max(max_row, v_max_row) + 1):
            for c in range(1, max(max_col, v_max_col) + 1):
                vc = ws_value.cell(row=r, column=c)
                if vc.value is not None:
                    value_lookup[(r, c)] = vc.value

    for row in ws_formula.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            r, c = cell.row, cell.column
            coord = cell.coordinate
            stats.cells_total += 1

            formula_val = cell.value
            evaluated_val = value_lookup.get((r, c))

            has_value = formula_val is not None or evaluated_val is not None
            is_formula = isinstance(formula_val, str) and formula_val.startswith("=")

            if not has_value and cell.comment is None and cell.hyperlink is None:
                # Truly empty cell with no metadata — skip to keep JSON small.
                continue

            entry: dict = {
                "row": r,
                "col": c,
                "col_letter": get_column_letter(c),
                "data_type": cell.data_type,
                "number_format": cell.number_format,
            }

            if has_value:
                stats.cells_with_value += 1
                entry["value"] = _safe_serialise(evaluated_val if evaluated_val is not None else formula_val)

            if is_formula:
                stats.cells_with_formula += 1
                entry["formula"] = formula_val
                if evaluated_val is not None and evaluated_val != formula_val:
                    entry["computed"] = _safe_serialise(evaluated_val)

            if cell.comment is not None:
                stats.cells_with_comment += 1
                entry["comment"] = {
                    "text": cell.comment.text,
                    "author": cell.comment.author,
                }

            if cell.hyperlink is not None:
                stats.cells_with_hyperlink += 1
                try:
                    entry["hyperlink"] = {
                        "target": cell.hyperlink.target,
                        "tooltip": cell.hyperlink.tooltip,
                        "display": cell.hyperlink.display,
                    }
                except Exception:
                    entry["hyperlink"] = {"raw": str(cell.hyperlink)}

            # Style — deduplicate.
            font_d = _font_to_dict(cell.font)
            fill_d = _fill_to_dict(cell.fill)
            border_d = _border_to_dict(cell.border)
            align_d = _alignment_to_dict(cell.alignment)
            sig = _style_signature(font_d, fill_d, border_d, align_d, cell.number_format)
            if sig not in style_table:
                style_table[sig] = len(style_table)
            cell_styles[coord] = style_table[sig]

            cells[coord] = entry

    stats.style_records = len(style_table)

    # Emit cells.json
    with open(sheet_dir / "cells.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws_formula.title, "cells": cells}, f, ensure_ascii=False, indent=1, default=str)

    # Emit styles.json — inverted style table + per-cell style id map.
    inv_styles: dict[int, dict] = {}
    for sig, sid in style_table.items():
        inv_styles[sid] = json.loads(sig)
    with open(sheet_dir / "styles.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws_formula.title, "styles": inv_styles, "cell_to_style": cell_styles}, f, ensure_ascii=False, indent=1, default=str)


def _extract_sheet_layout(ws: Worksheet, sheet_dir: Path, stats: SheetStats) -> None:
    """Merged ranges, frozen panes, hidden rows/cols, dimensions, print areas."""
    merged = [str(mr) for mr in ws.merged_cells.ranges]
    stats.merged_count = len(merged)

    hidden_rows = []
    for row_idx, dim in ws.row_dimensions.items():
        if getattr(dim, "hidden", False):
            hidden_rows.append(row_idx)
    stats.hidden_rows = len(hidden_rows)

    hidden_cols = []
    col_widths: dict[str, float] = {}
    for col_letter, dim in ws.column_dimensions.items():
        if getattr(dim, "hidden", False):
            hidden_cols.append(col_letter)
        if dim.width is not None:
            col_widths[col_letter] = float(dim.width)
    stats.hidden_cols = len(hidden_cols)

    row_heights: dict[int, float] = {}
    for row_idx, dim in ws.row_dimensions.items():
        if dim.height is not None:
            row_heights[row_idx] = float(dim.height)

    layout = {
        "sheet": ws.title,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_ranges": merged,
        "frozen_panes": ws.freeze_panes,
        "hidden_rows": hidden_rows,
        "hidden_cols": hidden_cols,
        "col_widths": col_widths,
        "row_heights": row_heights,
        "print_area": ws.print_area,
        "print_titles": {
            "rows": ws.print_title_rows,
            "cols": ws.print_title_cols,
        },
        "sheet_state": ws.sheet_state,  # "visible" / "hidden" / "veryHidden"
        "tab_color": _color_to_str(getattr(ws.sheet_properties, "tabColor", None)),
    }
    with open(sheet_dir / "layout.json", "w", encoding="utf-8") as f:
        json.dump(layout, f, ensure_ascii=False, indent=2, default=str)


def _extract_sheet_tables(ws: Worksheet, sheet_dir: Path, stats: SheetStats) -> None:
    """Excel structured tables and autofilter."""
    tables = []
    try:
        for tbl_name, tbl in ws.tables.items():
            tables.append({
                "name": tbl_name,
                "ref": tbl.ref if hasattr(tbl, "ref") else None,
                "display_name": getattr(tbl, "displayName", None),
                "header_row_count": getattr(tbl, "headerRowCount", None),
                "totals_row_count": getattr(tbl, "totalsRowCount", None),
            })
    except Exception:
        pass
    stats.table_count = len(tables)

    autofilter = None
    if ws.auto_filter and ws.auto_filter.ref:
        autofilter = {"ref": ws.auto_filter.ref}

    with open(sheet_dir / "tables.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws.title, "tables": tables, "autofilter": autofilter}, f, ensure_ascii=False, indent=2, default=str)


def _extract_sheet_charts(ws: Worksheet, sheet_dir: Path, stats: SheetStats) -> None:
    """Chart kinds + series + ranges."""
    charts = []
    for ch in ws._charts:
        try:
            charts.append({
                "type": type(ch).__name__,
                "title": str(ch.title) if ch.title else None,
                "series": [str(s) for s in getattr(ch, "series", [])],
                "anchor": str(ch.anchor) if hasattr(ch, "anchor") else None,
            })
        except Exception as e:
            charts.append({"type": type(ch).__name__, "error": str(e)})
    stats.chart_count = len(charts)
    with open(sheet_dir / "charts.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws.title, "charts": charts}, f, ensure_ascii=False, indent=2, default=str)


def _extract_sheet_validation(ws: Worksheet, sheet_dir: Path, stats: SheetStats) -> None:
    """Data validation rules (dropdowns, ranges, allowed values)."""
    validations = []
    try:
        for dv in ws.data_validations.dataValidation:
            validations.append({
                "type": dv.type,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "operator": dv.operator,
                "allowBlank": dv.allowBlank,
                "showDropDown": dv.showDropDown,
                "showInputMessage": dv.showInputMessage,
                "showErrorMessage": dv.showErrorMessage,
                "errorTitle": dv.errorTitle,
                "error": dv.error,
                "sqref": str(dv.sqref) if dv.sqref else None,
            })
    except Exception:
        pass
    stats.validation_count = len(validations)
    with open(sheet_dir / "validation.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws.title, "validations": validations}, f, ensure_ascii=False, indent=2, default=str)


def _extract_conditional_formats(ws: Worksheet, sheet_dir: Path) -> None:
    """Conditional formatting rules per range."""
    cfs = []
    try:
        for rng, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                cfs.append({
                    "range": str(rng),
                    "type": rule.type,
                    "priority": rule.priority,
                    "formula": list(rule.formula) if rule.formula else None,
                    "operator": rule.operator,
                    "text": rule.text,
                    "stopIfTrue": rule.stopIfTrue,
                })
    except Exception:
        pass
    with open(sheet_dir / "conditional_formats.json", "w", encoding="utf-8") as f:
        json.dump({"sheet": ws.title, "rules": cfs}, f, ensure_ascii=False, indent=2, default=str)


# ─── Workbook-level extractor ─────────────────────────────────────────────


def _extract_workbook_meta(wb, out_dir: Path) -> dict:
    """Defined names, props, security, sheet order."""
    defined_names = []
    try:
        for dn in wb.defined_names.definedName:
            defined_names.append({
                "name": dn.name,
                "value": dn.value,
                "comment": dn.comment,
                "hidden": dn.hidden,
                "localSheetId": dn.localSheetId,
            })
    except Exception:
        # openpyxl ≥3.1 changes API.
        try:
            for name in wb.defined_names:
                dn = wb.defined_names[name]
                defined_names.append({"name": name, "value": getattr(dn, "value", None)})
        except Exception:
            pass

    props = {}
    try:
        p = wb.properties
        props = {
            "creator": p.creator,
            "title": p.title,
            "subject": p.subject,
            "description": p.description,
            "keywords": p.keywords,
            "lastModifiedBy": p.lastModifiedBy,
            "category": p.category,
            "created": p.created.isoformat() if p.created else None,
            "modified": p.modified.isoformat() if p.modified else None,
            "company": getattr(p, "company", None),
            "manager": getattr(p, "manager", None),
        }
    except Exception:
        pass

    meta = {
        "sheet_names": wb.sheetnames,
        "active_sheet": wb.active.title if wb.active else None,
        "defined_names": defined_names,
        "properties": props,
        "custom_doc_props": {},
    }

    # Custom doc props.
    try:
        for cp in wb.custom_doc_props.props:
            meta["custom_doc_props"][cp.name] = {
                "type": cp.type,
                "value": _safe_serialise(cp.value),
            }
    except Exception:
        pass

    with open(out_dir / "workbook.json", "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2, default=str)
    return meta


# ─── Top-level entry ──────────────────────────────────────────────────────


def extract_workbook(xlsx_path: str | Path, out_dir: str | Path) -> ExtractionStats:
    """Run the full extraction. Returns aggregate stats.

    No raw cell values are returned — only counts. The caller (CLI / harness)
    is responsible for never logging the raw JSON to STDOUT; everything stays
    on disk under `out_dir`.
    """
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Pass 1: formulas preserved.
    wb_f = load_workbook(xlsx_path, data_only=False, read_only=False)
    # Pass 2: cached evaluated values.
    wb_v = load_workbook(xlsx_path, data_only=True, read_only=False)

    meta = _extract_workbook_meta(wb_f, out_dir)
    sheet_stats: list[SheetStats] = []

    for sheet_name in wb_f.sheetnames:
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name] if sheet_name in wb_v.sheetnames else None
        sheet_dir = out_dir / "sheets" / sheet_name
        sheet_dir.mkdir(parents=True, exist_ok=True)
        stats = SheetStats(
            name=sheet_name,
            max_row=ws_f.max_row or 0,
            max_col=ws_f.max_column or 0,
        )
        _extract_sheet_cells(ws_f, ws_v, sheet_dir, stats)
        _extract_sheet_layout(ws_f, sheet_dir, stats)
        _extract_sheet_tables(ws_f, sheet_dir, stats)
        _extract_sheet_charts(ws_f, sheet_dir, stats)
        _extract_sheet_validation(ws_f, sheet_dir, stats)
        _extract_conditional_formats(ws_f, sheet_dir)
        sheet_stats.append(stats)

    agg = ExtractionStats(
        workbook=str(xlsx_path.name),
        sheet_count=len(wb_f.sheetnames),
        defined_names=len(meta.get("defined_names", [])),
        total_cells=sum(s.cells_total for s in sheet_stats),
        total_formulas=sum(s.cells_with_formula for s in sheet_stats),
        total_comments=sum(s.cells_with_comment for s in sheet_stats),
        total_hyperlinks=sum(s.cells_with_hyperlink for s in sheet_stats),
        out_dir=str(out_dir.resolve()),
        sheets=sheet_stats,
    )

    with open(out_dir / "extraction_summary.json", "w", encoding="utf-8") as f:
        json.dump({**asdict(agg), "sheets": [asdict(s) for s in sheet_stats]}, f, ensure_ascii=False, indent=2, default=str)

    return agg
